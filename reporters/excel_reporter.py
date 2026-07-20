from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelReporter:
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    severity_fills = {
        "CRITICAL": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "HIGH": PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
        "MEDIUM": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
        "LOW": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    }

    def generate(self, results: list, name: str = "vulnerability_report") -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        headers = ["Repository", "Scanner", "Critical", "High", "Medium", "Low", "Total", "Errors"]
        ws.append(headers)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

        for row_idx, r in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=r.repo).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=r.scanner).border = self.thin_border
            s = r.summary
            ws.cell(row=row_idx, column=3, value=s["CRITICAL"]).border = self.thin_border
            ws.cell(row=row_idx, column=4, value=s["HIGH"]).border = self.thin_border
            ws.cell(row=row_idx, column=5, value=s["MEDIUM"]).border = self.thin_border
            ws.cell(row=row_idx, column=6, value=s["LOW"]).border = self.thin_border
            ws.cell(row=row_idx, column=7, value=len(r.vulnerabilities)).border = self.thin_border
            ws.cell(row=row_idx, column=8, value="; ".join(r.errors)).border = self.thin_border

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 12
        for col in "CDEFGH":
            ws.column_dimensions[col].width = 10

        details_ws = wb.create_sheet("Details")
        detail_headers = ["Repository", "Scanner", "Vulnerability ID", "Package", "Installed", "Fixed", "Severity", "Type", "Description"]
        details_ws.append(detail_headers)
        for col_idx, header in enumerate(detail_headers, 1):
            cell = details_ws.cell(row=1, column=col_idx)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

        row_idx = 2
        for r in results:
            for v in r.vulnerabilities:
                sev = v.severity.upper()
                details_ws.cell(row=row_idx, column=1, value=r.repo).border = self.thin_border
                details_ws.cell(row=row_idx, column=2, value=r.scanner).border = self.thin_border
                details_ws.cell(row=row_idx, column=3, value=v.id).border = self.thin_border
                details_ws.cell(row=row_idx, column=4, value=v.package).border = self.thin_border
                details_ws.cell(row=row_idx, column=5, value=v.installed_version).border = self.thin_border
                details_ws.cell(row=row_idx, column=6, value=v.fixed_version or "N/A").border = self.thin_border
                sev_cell = details_ws.cell(row=row_idx, column=7, value=sev)
                sev_cell.border = self.thin_border
                if sev in self.severity_fills:
                    sev_cell.fill = self.severity_fills[sev]
                details_ws.cell(row=row_idx, column=8, value=v.type).border = self.thin_border
                details_ws.cell(row=row_idx, column=9, value=v.description).border = self.thin_border
                row_idx += 1

        detail_widths = [40, 12, 20, 20, 15, 15, 12, 15, 50]
        for i, w in enumerate(detail_widths, 1):
            details_ws.column_dimensions[get_column_letter(i)].width = w

        out_path = self.output_dir / f"{name}.xlsx"
        wb.save(str(out_path))
        return str(out_path)
